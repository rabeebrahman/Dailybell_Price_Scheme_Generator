import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def process_files(quote_file, slab_file, output_file):
    df = pd.read_excel(quote_file)
    slabs_df = pd.read_excel(slab_file)

    df.columns = df.columns.str.strip()
    slabs_df.columns = slabs_df.columns.str.strip()
    df["Product Name"] = df["Product Name"].astype(str).str.strip().str.lower()
    slabs_df["Product Name"] = slabs_df["Product Name"].astype(str).str.strip().str.lower()
    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce")

    df["Scheme Price"] = np.nan
    df["Reason"] = ""

    for i, row in df.iterrows():
        product = row["Product Name"]
        qty = row["Qty"]

        if pd.isna(product) or pd.isna(qty):
            continue

        matching_slabs = slabs_df[slabs_df["Product Name"] == product]
        if matching_slabs.empty:
            continue

        for _, slab in matching_slabs.iterrows():
            cond_type = str(slab["Condition Type"]).strip().lower()
            min_qty = slab["Min Qty"]
            max_qty = slab["Max Qty"]
            price = slab["Price per Unit"]

            if pd.isna(price):
                continue

            matched = False
            if cond_type == "lt" and qty < min_qty:
                matched = True
            elif cond_type == "lte" and qty <= min_qty:
                matched = True
            elif cond_type == "gte" and qty >= min_qty:
                matched = True
            elif cond_type == "equals" and qty == min_qty:
                matched = True
            elif cond_type == "range" and pd.notna(min_qty) and pd.notna(max_qty):
                if min_qty <= qty <= max_qty:
                    matched = True

            if matched:
                df.at[i, "Scheme Price"] = round(price, 2)
                if cond_type == "range":
                    reason = f"range slab {int(min_qty)}–{int(max_qty)}"
                elif cond_type == "equals":
                    reason = f"equals {int(min_qty)}"
                elif cond_type == "gte":
                    reason = f">= {int(min_qty)}"
                elif cond_type == "lt":
                    reason = f"< {int(min_qty)}"
                elif cond_type == "lte":
                    reason = f"<= {int(min_qty)}"
                else:
                    reason = "matched slab"
                df.at[i, "Reason"] = reason
                break

    df["Scheme Price"] = df["Scheme Price"].fillna(0)
    df["Reason"] = df["Reason"].replace("", "No Match")
    df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active
    header = [cell.value for cell in ws[1]]

    price_col = header.index("Price with Tax") + 1
    scheme_col = header.index("Scheme Price") + 1
    reason_col = header.index("Reason") + 1

    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        reason_val = ws.cell(row=row, column=reason_col).value
        price_val = ws.cell(row=row, column=price_col).value
        scheme_val = ws.cell(row=row, column=scheme_col).value

        if reason_val == "No Match":
            continue

        if isinstance(price_val, (int, float)) and isinstance(scheme_val, (int, float)):
            if round(price_val, 2) != round(scheme_val, 2):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = highlight

    wb.save(output_file)