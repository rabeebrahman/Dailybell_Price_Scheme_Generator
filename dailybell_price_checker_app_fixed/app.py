import os
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from flask import Flask, render_template, request, send_file, redirect
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import threading
import time

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
PROCESSED_FOLDER = "processed"
SLAB_FILE = "final_price_slab_for_dailybell_retail_slab_generator.xlsx"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)

def delete_file_after_delay(filepath, delay=300):
    def delete():
        time.sleep(delay)
        if os.path.exists(filepath):
            os.remove(filepath)
    threading.Thread(target=delete).start()

def ensure_expected_columns(df):
    """
    Ensures that all expected columns exist.
    Missing columns are created with blank values.
    Extra columns are kept as-is.
    """
    expected_cols = [
        "Serial Number", "Quotation Date", "Product Name", "Qty", "Variant Name",
        "Item Total Amount", "Price with Tax", "Item Net Amount", "Tax (%)",
        "CGST", "IGST", "SGST", "HSN/SAC", "Party Name", "Party Company Name",
        "Party Mobile", "Party GSTIN", "Created By", "sales man"
    ]

    for col in expected_cols:
        if col not in df.columns:
            df[col] = np.nan  # create missing column
    return df

def apply_price_scheme(quotation_df, slabs_df):
    quotation_df["Product Name"] = quotation_df["Product Name"].astype(str).str.strip().str.lower()
    slabs_df["Product Name"] = slabs_df["Product Name"].astype(str).str.strip().str.lower()
    quotation_df["Qty"] = pd.to_numeric(quotation_df["Qty"], errors="coerce")

    quotation_df["Scheme Price"] = np.nan
    quotation_df["Reason"] = ""

    for i, row in quotation_df.iterrows():
        product = row["Product Name"]
        qty = row["Qty"]

        if pd.isna(product) or pd.isna(qty):
            continue

        matching_slabs = slabs_df[slabs_df["Product Name"] == product]

        if matching_slabs.empty:
            continue

        matched = False
        for _, slab in matching_slabs.iterrows():
            cond_type = str(slab["Condition Type"]).strip().lower()
            min_qty = slab["Min Qty"]
            max_qty = slab["Max Qty"]
            price = slab["Price per Unit"]

            if pd.isna(price):
                continue

            if cond_type == "lt" and qty < min_qty:
                matched = True
            elif cond_type == "lte" and qty <= min_qty:
                matched = True
            elif cond_type == "gte" and qty >= min_qty:
                matched = True
            elif cond_type == "equals" and qty == min_qty:
                matched = True
            elif cond_type == "range" and pd.notna(min_qty) and pd.notna(max_qty):
                if min_qty <= qty <= max_qty:
                    matched = True

            if matched:
                quotation_df.at[i, "Scheme Price"] = round(price, 2)
                if cond_type == "range":
                    reason = f"range slab {int(min_qty)}–{int(max_qty)}"
                elif cond_type == "equals":
                    reason = f"equals {int(min_qty)}"
                elif cond_type == "gte":
                    reason = f">= {int(min_qty)}"
                elif cond_type == "lt":
                    reason = f"< {int(min_qty)}"
                elif cond_type == "lte":
                    reason = f"<= {int(min_qty)}"
                else:
                    reason = "matched slab"
                quotation_df.at[i, "Reason"] = reason
                break

    quotation_df["Scheme Price"] = quotation_df["Scheme Price"].fillna(0)
    quotation_df["Reason"] = quotation_df["Reason"].replace("", "No Match")
    return quotation_df

def highlight_mismatches(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    price_with_tax_col = header.index("Price with Tax") + 1
    scheme_price_col = header.index("Scheme Price") + 1
    reason_col = header.index("Reason") + 1
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        tax_price = ws.cell(row=row, column=price_with_tax_col).value
        scheme_price = ws.cell(row=row, column=scheme_price_col).value
        reason = ws.cell(row=row, column=reason_col).value
        if isinstance(tax_price, (int, float)) and isinstance(scheme_price, (int, float)):
            if round(tax_price, 2) != round(scheme_price, 2) and reason != "No Match":
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = highlight_fill
    wb.save(excel_path)

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet

def generate_pdf(df, pdf_path):
    # Sort by Party Name if available
    if "Party Name" in df.columns:
        df = df.sort_values(by="Party Name", ascending=True)

    df = df.fillna("").astype(str)
    styles = getSampleStyleSheet()
    normal_style = styles["Normal"]

    # Wrap all text in Paragraph to allow line breaks
    data = [[Paragraph(str(col), normal_style) for col in df.columns]]
    for row in df.values.tolist():
        data.append([Paragraph(str(cell), normal_style) for cell in row])

    # Approximate equal column widths for the page
    col_count = len(df.columns)
    page_width = letter[0] - 40  # margin space
    col_widths = [page_width / col_count] * col_count

    doc = SimpleDocTemplate(pdf_path, pagesize=letter)
    elements = []

    # Split into chunks so the table doesn't exceed one page
    chunk_size = 40
    for start in range(0, len(data), chunk_size):
        chunk = data[start:start + chunk_size]
        table = Table(chunk, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.gray),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(table)
        elements.append(PageBreak())

    doc.build(elements)


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        uploaded_file = request.files["quotation_file"]
        if uploaded_file.filename == "":
            return redirect(request.url)

        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        input_path = os.path.join(UPLOAD_FOLDER, f"{timestamp}_{uploaded_file.filename}")
        uploaded_file.save(input_path)
        delete_file_after_delay(input_path)

        df_quotation = pd.read_excel(input_path)
        df_quotation = ensure_expected_columns(df_quotation)  # NEW: Fix missing cols
        df_slab = pd.read_excel(SLAB_FILE)
        df_result = apply_price_scheme(df_quotation, df_slab)

        xlsx_output = os.path.join(PROCESSED_FOLDER, f"{timestamp}_quotation_with_scheme.xlsx")
        df_result.to_excel(xlsx_output, index=False)
        highlight_mismatches(xlsx_output)

        pdf_output = os.path.join(PROCESSED_FOLDER, f"{timestamp}_quotation_with_scheme.pdf")
        generate_pdf(df_result, pdf_output)

        delete_file_after_delay(xlsx_output)
        delete_file_after_delay(pdf_output)

        return render_template(
            "index.html",
            xlsx_file=os.path.basename(xlsx_output),
            pdf_file=os.path.basename(pdf_output)
        )

    return render_template("index.html")

@app.route("/download/<filename>")
def download(filename):
    path = os.path.join(PROCESSED_FOLDER, filename)
    return send_file(path, as_attachment=True)

if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
